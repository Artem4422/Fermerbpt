import sqlite3
import database
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import qr_code


def generate_session_report_excel(session_id: int) -> io.BytesIO:
    """Генерирует полный Excel отчет по сессии"""
    session = database.get_session(session_id)
    if not session:
        raise ValueError("Сессия не найдена")
    
    # Получаем данные
    orders = database.get_session_orders(session_id)
    products = database.get_products_by_session(session_id)
    
    # Создаем рабочую книгу
    wb = Workbook()
    
    # Удаляем дефолтный лист
    wb.remove(wb.active)
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Лист 1: Общая информация
    ws_summary = wb.create_sheet("Общая информация")
    
    ws_summary['A1'] = f"ОТЧЕТ ПО СЕССИИ: {session['session_name']}"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A3'] = "Дата создания сессии:"
    ws_summary['B3'] = session['created_at']
    ws_summary['A4'] = "Всего заказов:"
    ws_summary['B4'] = len(orders)
    
    completed_orders = [o for o in orders if o['status'] == 'completed']
    pending_orders = [o for o in orders if o['status'] == 'pending']
    processing_orders = [o for o in orders if o['status'] == 'processing']
    cancelled_orders = [o for o in orders if o['status'] == 'cancelled']
    
    ws_summary['A5'] = "Выдано заказов:"
    ws_summary['B5'] = len(completed_orders)
    ws_summary['A6'] = "Ожидает обработки:"
    ws_summary['B6'] = len(pending_orders)
    ws_summary['A7'] = "В обработке:"
    ws_summary['B7'] = len(processing_orders)
    ws_summary['A8'] = "Отменено:"
    ws_summary['B8'] = len(cancelled_orders)
    
    total_revenue = sum(o['total_amount'] for o in completed_orders)
    ws_summary['A9'] = "Выручка (выданные заказы):"
    ws_summary['B9'] = total_revenue
    ws_summary['B9'].number_format = '#,##0.00₽'
    
    # Лист 2: Продажи
    ws_sales = wb.create_sheet("Продажи")
    
    headers_sales = ["№ заказа", "ФИО", "Телефон", "Товары", "Количество ящиков", "Сумма", "Статус", "Дата"]
    for col, header in enumerate(headers_sales, 1):
        cell = ws_sales.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    row = 2
    total_boxes_sold = 0
    for order in orders:
        order_items = database.get_order_items(order['order_id'])
        boxes_in_order = sum(item['quantity'] for item in order_items)
        total_boxes_sold += boxes_in_order if order['status'] == 'completed' else 0
        
        items_text = ", ".join([f"{item['product_name']} x{item['quantity']}" for item in order_items])
        
        ws_sales.cell(row=row, column=1).value = order['order_number']
        ws_sales.cell(row=row, column=2).value = order['full_name']
        ws_sales.cell(row=row, column=3).value = order['phone_number']
        ws_sales.cell(row=row, column=4).value = items_text
        ws_sales.cell(row=row, column=5).value = boxes_in_order
        ws_sales.cell(row=row, column=6).value = order['total_amount']
        ws_sales.cell(row=row, column=6).number_format = '#,##0.00₽'
        ws_sales.cell(row=row, column=7).value = database.get_order_status_ru(order['status'])
        ws_sales.cell(row=row, column=8).value = order['created_at']
        
        for col in range(1, 9):
            ws_sales.cell(row=row, column=col).border = border
        
        row += 1
    
    # Автоподбор ширины колонок
    for col in range(1, 9):
        ws_sales.column_dimensions[get_column_letter(col)].width = 15
    
    # Лист 3: Товары и остатки
    ws_products = wb.create_sheet("Товары и остатки")
    
    headers_products = ["Товар", "Цена за ящик", "Начальное количество", "Продано", "Остаток", "Выручка"]
    for col, header in enumerate(headers_products, 1):
        cell = ws_products.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    row = 2
    for product in products:
        # Подсчитываем проданное количество
        sold_count = 0
        product_revenue = 0
        
        for order in completed_orders:
            order_items = database.get_order_items(order['order_id'])
            for item in order_items:
                if item['product_id'] == product['product_id']:
                    sold_count += item['quantity']
                    product_revenue += item['quantity'] * item['price']
        
        remaining = product['boxes_count']
        
        ws_products.cell(row=row, column=1).value = product['product_name']
        ws_products.cell(row=row, column=2).value = product['price']
        ws_products.cell(row=row, column=2).number_format = '#,##0.00₽'
        ws_products.cell(row=row, column=3).value = product['boxes_count'] + sold_count  # Начальное = текущее + проданное
        ws_products.cell(row=row, column=4).value = sold_count
        ws_products.cell(row=row, column=5).value = remaining
        ws_products.cell(row=row, column=6).value = product_revenue
        ws_products.cell(row=row, column=6).number_format = '#,##0.00₽'
        
        for col in range(1, 7):
            ws_products.cell(row=row, column=col).border = border
        
        row += 1
    
    # Автоподбор ширины колонок
    for col in range(1, 7):
        ws_products.column_dimensions[get_column_letter(col)].width = 20
    
    # Лист 4: Клиенты (отсортированные по количеству заказов)
    ws_customers = wb.create_sheet("Клиенты")
    
    headers_customers = ["№", "ФИО", "Телефон", "Количество заказов", "Номера заказов", "Общая сумма", "Всего ящиков"]
    for col, header in enumerate(headers_customers, 1):
        cell = ws_customers.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Группируем заказы по клиентам
    customers_dict = {}
    for order in orders:
        # Используем комбинацию ФИО и телефона как ключ
        customer_key = f"{order['full_name']}_{order['phone_number']}"
        
        if customer_key not in customers_dict:
            customers_dict[customer_key] = {
                'full_name': order['full_name'],
                'phone_number': order['phone_number'],
                'user_id': order['user_id'],
                'orders': [],
                'total_amount': 0,
                'total_boxes': 0
            }
        
        customers_dict[customer_key]['orders'].append(order)
        customers_dict[customer_key]['total_amount'] += order['total_amount']
        
        # Подсчитываем ящики в заказе
        order_items = database.get_order_items(order['order_id'])
        boxes_in_order = sum(item['quantity'] for item in order_items)
        customers_dict[customer_key]['total_boxes'] += boxes_in_order
    
    # Сортируем клиентов по количеству заказов (от большего к меньшему)
    customers_list = sorted(
        customers_dict.values(),
        key=lambda x: len(x['orders']),
        reverse=True
    )
    
    row = 2
    for idx, customer in enumerate(customers_list, 1):
        order_numbers = ", ".join([f"#{o['order_number']}" for o in customer['orders']])
        
        ws_customers.cell(row=row, column=1).value = idx
        ws_customers.cell(row=row, column=2).value = customer['full_name']
        ws_customers.cell(row=row, column=3).value = customer['phone_number']
        ws_customers.cell(row=row, column=4).value = len(customer['orders'])
        ws_customers.cell(row=row, column=5).value = order_numbers
        ws_customers.cell(row=row, column=6).value = customer['total_amount']
        ws_customers.cell(row=row, column=6).number_format = '#,##0.00₽'
        ws_customers.cell(row=row, column=7).value = customer['total_boxes']
        
        for col in range(1, 8):
            ws_customers.cell(row=row, column=col).border = border
        
        row += 1
    
    # Автоподбор ширины колонок
    ws_customers.column_dimensions['A'].width = 8
    ws_customers.column_dimensions['B'].width = 25
    ws_customers.column_dimensions['C'].width = 15
    ws_customers.column_dimensions['D'].width = 18
    ws_customers.column_dimensions['E'].width = 50
    ws_customers.column_dimensions['F'].width = 15
    ws_customers.column_dimensions['G'].width = 15
    
    # Лист 5: Статистика
    ws_stats = wb.create_sheet("Статистика")
    
    ws_stats['A1'] = "СТАТИСТИКА ПО СЕССИИ"
    ws_stats['A1'].font = title_font
    ws_stats.merge_cells('A1:B1')
    
    row = 3
    stats = [
        ("Всего заказов", len(orders)),
        ("Выдано заказов", len(completed_orders)),
        ("Ожидает обработки", len(pending_orders)),
        ("В обработке", len(processing_orders)),
        ("Отменено", len(cancelled_orders)),
        ("", ""),
        ("Всего клиентов", len(customers_list)),
        ("", ""),
        ("Общая выручка", total_revenue),
        ("Всего продано ящиков", total_boxes_sold),
        ("Средний чек", total_revenue / len(completed_orders) if completed_orders else 0),
    ]
    
    for label, value in stats:
        ws_stats.cell(row=row, column=1).value = label
        ws_stats.cell(row=row, column=1).font = Font(bold=True)
        if isinstance(value, (int, float)) and value != 0:
            ws_stats.cell(row=row, column=2).value = value
            if 'выручка' in label.lower() or 'чек' in label.lower():
                ws_stats.cell(row=row, column=2).number_format = '#,##0.00₽'
        else:
            ws_stats.cell(row=row, column=2).value = value
        row += 1
    
    # Сохраняем в байты
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    return excel_bytes


def generate_period_report_excel(period: str) -> io.BytesIO:
    """Генерирует полный Excel отчет за период по всем сессиям"""
    from datetime import datetime, timedelta
    
    # Получаем данные за период
    orders = database.get_orders_by_period(period)
    
    # Получаем все сессии, которые были в этом периоде
    session_ids = list(set([o['session_id'] for o in orders]))
    sessions = [database.get_session(sid) for sid in session_ids if database.get_session(sid)]
    
    # Определяем название периода
    period_names = {
        "week": "неделю",
        "month": "месяц",
        "year": "год",
        "all_time": "все время"
    }
    period_name = period_names.get(period, period)
    
    # Создаем рабочую книгу
    wb = Workbook()
    wb.remove(wb.active)
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Лист 1: Общая информация
    ws_summary = wb.create_sheet("Общая информация")
    
    ws_summary['A1'] = f"ОТЧЕТ ЗА {period_name.upper()}"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    
    completed_orders = [o for o in orders if o['status'] == 'completed']
    pending_orders = [o for o in orders if o['status'] == 'pending']
    processing_orders = [o for o in orders if o['status'] == 'processing']
    cancelled_orders = [o for o in orders if o['status'] == 'cancelled']
    
    total_revenue = sum(o['total_amount'] for o in completed_orders)
    total_boxes_sold = 0
    for order in completed_orders:
        order_items = database.get_order_items(order['order_id'])
        total_boxes_sold += sum(item['quantity'] for item in order_items)
    
    ws_summary['A3'] = "Период:"
    ws_summary['B3'] = period_name
    ws_summary['A4'] = "Всего заказов:"
    ws_summary['B4'] = len(orders)
    ws_summary['A5'] = "Выдано заказов:"
    ws_summary['B5'] = len(completed_orders)
    ws_summary['A6'] = "Ожидает обработки:"
    ws_summary['B6'] = len(pending_orders)
    ws_summary['A7'] = "В обработке:"
    ws_summary['B7'] = len(processing_orders)
    ws_summary['A8'] = "Отменено:"
    ws_summary['B8'] = len(cancelled_orders)
    ws_summary['A9'] = "Всего сессий:"
    ws_summary['B9'] = len(sessions)
    ws_summary['A10'] = "Общая выручка:"
    ws_summary['B10'] = total_revenue
    ws_summary['B10'].number_format = '#,##0.00₽'
    ws_summary['A11'] = "Всего продано ящиков:"
    ws_summary['B11'] = total_boxes_sold
    
    # Лист 2: Все заказы
    ws_orders = wb.create_sheet("Все заказы")
    
    headers_orders = ["№ заказа", "Сессия", "ФИО", "Телефон", "Товары", "Ящиков", "Сумма", "Статус", "Дата"]
    
    # Если период "all_time", группируем по сессиям с заголовками
    if period == "all_time":
        row = 1
        
        # Группируем заказы по сессиям
        orders_by_session = {}
        for order in orders:
            session_id = order['session_id']
            if session_id not in orders_by_session:
                orders_by_session[session_id] = []
            orders_by_session[session_id].append(order)
        
        # Сортируем сессии по дате создания (если есть)
        sorted_sessions = sorted(sessions, key=lambda s: s.get('created_at', ''), reverse=True)
        
        session_header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        session_header_font = Font(bold=True, size=13)
        
        for idx, session in enumerate(sorted_sessions):
            session_id = session['session_id']
            if session_id not in orders_by_session:
                continue
            
            session_orders = orders_by_session[session_id]
            
            # Заголовок сессии (для каждой сессии, включая первую)
            ws_orders.cell(row=row, column=1).value = f"СЕССИЯ: {session['session_name']}"
            ws_orders.cell(row=row, column=1).font = session_header_font
            ws_orders.cell(row=row, column=1).fill = session_header_fill
            ws_orders.merge_cells(f'A{row}:I{row}')
            row += 1
            
            # Заголовки таблицы для этой сессии (после каждого заголовка сессии)
            for col, header in enumerate(headers_orders, 1):
                cell = ws_orders.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            row += 1
            
            # Заказы этой сессии
            session_total_amount = 0
            session_total_boxes = 0
            session_completed_count = 0
            last_order_date = None
            
            for order in session_orders:
                order_items = database.get_order_items(order['order_id'])
                boxes_in_order = sum(item['quantity'] for item in order_items)
                
                ws_orders.cell(row=row, column=1).value = order['order_number']
                ws_orders.cell(row=row, column=2).value = session['session_name']
                ws_orders.cell(row=row, column=3).value = order['full_name']
                ws_orders.cell(row=row, column=4).value = order['phone_number']
                ws_orders.cell(row=row, column=5).value = order['items']
                ws_orders.cell(row=row, column=6).value = boxes_in_order
                ws_orders.cell(row=row, column=7).value = order['total_amount']
                ws_orders.cell(row=row, column=7).number_format = '#,##0.00₽'
                ws_orders.cell(row=row, column=8).value = database.get_order_status_ru(order['status'])
                ws_orders.cell(row=row, column=9).value = order['created_at']
                
                # Подсчитываем статистику
                session_total_amount += order['total_amount']
                session_total_boxes += boxes_in_order
                if order['status'] == 'completed':
                    session_completed_count += 1
                if not last_order_date or order['created_at'] > last_order_date:
                    last_order_date = order['created_at']
                
                for col in range(1, 10):
                    ws_orders.cell(row=row, column=col).border = border
                
                row += 1
            
            # Добавляем строку с итогами сессии
            summary_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            summary_font = Font(bold=True, size=11)
            
            # Статус сессии
            session_status = "Активна" if session.get('is_active') else "Закрыта"
            # Дата завершения (дата последнего заказа или дата создания сессии)
            session_end_date = last_order_date if last_order_date else session.get('created_at', '')
            
            ws_orders.cell(row=row, column=1).value = "ИТОГИ СЕССИИ:"
            ws_orders.cell(row=row, column=1).font = summary_font
            ws_orders.cell(row=row, column=1).fill = summary_fill
            ws_orders.cell(row=row, column=2).value = f"Заказов: {len(session_orders)}"
            ws_orders.cell(row=row, column=2).font = summary_font
            ws_orders.cell(row=row, column=2).fill = summary_fill
            ws_orders.cell(row=row, column=3).value = f"Выдано: {session_completed_count}"
            ws_orders.cell(row=row, column=3).font = summary_font
            ws_orders.cell(row=row, column=3).fill = summary_fill
            ws_orders.cell(row=row, column=4).value = f"Ящиков: {session_total_boxes}"
            ws_orders.cell(row=row, column=4).font = summary_font
            ws_orders.cell(row=row, column=4).fill = summary_fill
            ws_orders.cell(row=row, column=5).value = f"Выручка: {session_total_amount:.2f}₽"
            ws_orders.cell(row=row, column=5).font = summary_font
            ws_orders.cell(row=row, column=5).fill = summary_fill
            ws_orders.cell(row=row, column=6).value = ""  # Пусто
            ws_orders.cell(row=row, column=6).fill = summary_fill
            ws_orders.cell(row=row, column=7).value = ""  # Пусто
            ws_orders.cell(row=row, column=7).fill = summary_fill
            ws_orders.cell(row=row, column=8).value = f"Статус: {session_status}"
            ws_orders.cell(row=row, column=8).font = summary_font
            ws_orders.cell(row=row, column=8).fill = summary_fill
            ws_orders.cell(row=row, column=9).value = f"Дата: {session_end_date}"
            ws_orders.cell(row=row, column=9).font = summary_font
            ws_orders.cell(row=row, column=9).fill = summary_fill
            
            # Применяем границы и объединяем ячейки для лучшего вида
            for col in range(1, 10):
                ws_orders.cell(row=row, column=col).border = border
            
            row += 1
    else:
        # Для других периодов оставляем как было
        for col, header in enumerate(headers_orders, 1):
            cell = ws_orders.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        row = 2
        for order in orders:
            session = database.get_session(order['session_id'])
            session_name = session['session_name'] if session else f"Сессия {order['session_id']}"
            order_items = database.get_order_items(order['order_id'])
            boxes_in_order = sum(item['quantity'] for item in order_items)
            
            ws_orders.cell(row=row, column=1).value = order['order_number']
            ws_orders.cell(row=row, column=2).value = session_name
            ws_orders.cell(row=row, column=3).value = order['full_name']
            ws_orders.cell(row=row, column=4).value = order['phone_number']
            ws_orders.cell(row=row, column=5).value = order['items']
            ws_orders.cell(row=row, column=6).value = boxes_in_order
            ws_orders.cell(row=row, column=7).value = order['total_amount']
            ws_orders.cell(row=row, column=7).number_format = '#,##0.00₽'
            ws_orders.cell(row=row, column=8).value = database.get_order_status_ru(order['status'])
            ws_orders.cell(row=row, column=9).value = order['created_at']
            
            for col in range(1, 10):
                ws_orders.cell(row=row, column=col).border = border
            
            row += 1
    
    # Автоподбор ширины колонок
    for col in range(1, 10):
        ws_orders.column_dimensions[get_column_letter(col)].width = 15
    
    # Лист 3: Клиенты (отсортированные по количеству заказов)
    ws_customers = wb.create_sheet("Клиенты")
    
    headers_customers = ["№", "ФИО", "Телефон", "Количество заказов", "Номера заказов", "Общая сумма", "Всего ящиков"]
    for col, header in enumerate(headers_customers, 1):
        cell = ws_customers.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Группируем заказы по клиентам
    customers_dict = {}
    for order in orders:
        customer_key = f"{order['full_name']}_{order['phone_number']}"
        
        if customer_key not in customers_dict:
            customers_dict[customer_key] = {
                'full_name': order['full_name'],
                'phone_number': order['phone_number'],
                'orders': [],
                'total_amount': 0,
                'total_boxes': 0
            }
        
        customers_dict[customer_key]['orders'].append(order)
        customers_dict[customer_key]['total_amount'] += order['total_amount']
        
        order_items = database.get_order_items(order['order_id'])
        boxes_in_order = sum(item['quantity'] for item in order_items)
        customers_dict[customer_key]['total_boxes'] += boxes_in_order
    
    # Сортируем клиентов по количеству заказов
    customers_list = sorted(
        customers_dict.values(),
        key=lambda x: len(x['orders']),
        reverse=True
    )
    
    row = 2
    for idx, customer in enumerate(customers_list, 1):
        order_numbers = ", ".join([f"#{o['order_number']}" for o in customer['orders']])
        
        ws_customers.cell(row=row, column=1).value = idx
        ws_customers.cell(row=row, column=2).value = customer['full_name']
        ws_customers.cell(row=row, column=3).value = customer['phone_number']
        ws_customers.cell(row=row, column=4).value = len(customer['orders'])
        ws_customers.cell(row=row, column=5).value = order_numbers
        ws_customers.cell(row=row, column=6).value = customer['total_amount']
        ws_customers.cell(row=row, column=6).number_format = '#,##0.00₽'
        ws_customers.cell(row=row, column=7).value = customer['total_boxes']
        
        for col in range(1, 8):
            ws_customers.cell(row=row, column=col).border = border
        
        row += 1
    
    # Автоподбор ширины колонок
    ws_customers.column_dimensions['A'].width = 8
    ws_customers.column_dimensions['B'].width = 25
    ws_customers.column_dimensions['C'].width = 15
    ws_customers.column_dimensions['D'].width = 18
    ws_customers.column_dimensions['E'].width = 50
    ws_customers.column_dimensions['F'].width = 15
    ws_customers.column_dimensions['G'].width = 15
    
    # Лист 4: Статистика по сессиям
    ws_sessions = wb.create_sheet("Статистика по сессиям")
    
    headers_sessions = ["Сессия", "Заказов", "Выдано", "Выручка", "Ящиков продано"]
    for col, header in enumerate(headers_sessions, 1):
        cell = ws_sessions.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    row = 2
    for session in sessions:
        session_orders = [o for o in orders if o['session_id'] == session['session_id']]
        session_completed = [o for o in session_orders if o['status'] == 'completed']
        session_revenue = sum(o['total_amount'] for o in session_completed)
        session_boxes = 0
        for order in session_completed:
            order_items = database.get_order_items(order['order_id'])
            session_boxes += sum(item['quantity'] for item in order_items)
        
        ws_sessions.cell(row=row, column=1).value = session['session_name']
        ws_sessions.cell(row=row, column=2).value = len(session_orders)
        ws_sessions.cell(row=row, column=3).value = len(session_completed)
        ws_sessions.cell(row=row, column=4).value = session_revenue
        ws_sessions.cell(row=row, column=4).number_format = '#,##0.00₽'
        ws_sessions.cell(row=row, column=5).value = session_boxes
        
        for col in range(1, 6):
            ws_sessions.cell(row=row, column=col).border = border
        
        row += 1
    
    # Автоподбор ширины колонок
    for col in range(1, 6):
        ws_sessions.column_dimensions[get_column_letter(col)].width = 20
    
    # Сохраняем в байты
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    return excel_bytes


def generate_channel_report_excel(session_id: int) -> io.BytesIO:
    """Генерирует Excel отчет для канала с маскировкой данных в формате двух столбцов"""
    session = database.get_session(session_id)
    if not session:
        raise ValueError("Сессия не найдена")
    
    # Получаем данные
    orders = database.get_session_orders(session_id)
    
    # Создаем рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет для канала"
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Заголовки для первого столбца (колонки A-F)
    headers_left = ["н", "тел", "фио", "тов", "я", "сум"]
    for col, header in enumerate(headers_left, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Заголовки для второго столбца (колонки H-M)
    headers_right = ["н", "тел", "фио", "тов", "я", "сум"]
    for col, header in enumerate(headers_right, 1):
        cell = ws.cell(row=1, column=col + 7)  # +7 чтобы начать с колонки H (8-я колонка)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Разделяем заказы на две части для двух столбцов
    total_orders = len(orders)
    mid_point = (total_orders + 1) // 2  # Округляем вверх
    
    orders_left = orders[:mid_point]
    orders_right = orders[mid_point:]
    
    # Заполняем левый столбец
    row = 2
    row_number = 1
    for order in orders_left:
        # Маскируем данные
        masked_name = qr_code.mask_name_channel(order['full_name'])
        masked_phone = qr_code.mask_phone_channel(order['phone_number'])
        
        # Получаем товары заказа
        order_items = database.get_order_items(order['order_id'])
        if order_items:
            # Берем первый товар (или объединяем если несколько)
            product_names = [item['product_name'] for item in order_items]
            product_name = ", ".join(set(product_names))  # Убираем дубликаты
            # Обрезаем до 9 символов
            product_name = product_name[:9] if len(product_name) > 9 else product_name
            total_quantity = sum(item['quantity'] for item in order_items)
        else:
            product_name = "Нет товаров"
            total_quantity = 0
        
        # Заполняем данные
        ws.cell(row=row, column=1).value = row_number  # Номер строки
        ws.cell(row=row, column=2).value = order['order_number']
        ws.cell(row=row, column=3).value = masked_phone
        ws.cell(row=row, column=4).value = masked_name
        ws.cell(row=row, column=5).value = product_name
        ws.cell(row=row, column=6).value = total_quantity
        ws.cell(row=row, column=7).value = int(order['total_amount'])
        ws.cell(row=row, column=7).number_format = '#,##0'
        
        # Применяем границы
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        
        row += 1
        row_number += 1
    
    # Заполняем правый столбец (продолжаем нумерацию с левой таблицы)
    row = 2
    # Продолжаем нумерацию с того места, где закончилась левая таблица
    row_number = len(orders_left) + 1
    for order in orders_right:
        # Маскируем данные
        masked_name = qr_code.mask_name_channel(order['full_name'])
        masked_phone = qr_code.mask_phone_channel(order['phone_number'])
        
        # Получаем товары заказа
        order_items = database.get_order_items(order['order_id'])
        if order_items:
            # Берем первый товар (или объединяем если несколько)
            product_names = [item['product_name'] for item in order_items]
            product_name = ", ".join(set(product_names))  # Убираем дубликаты
            # Обрезаем до 9 символов
            product_name = product_name[:9] if len(product_name) > 9 else product_name
            total_quantity = sum(item['quantity'] for item in order_items)
        else:
            product_name = "Нет товаров"
            total_quantity = 0
        
        # Заполняем данные (начинаем с колонки I, т.е. колонка 9)
        ws.cell(row=row, column=9).value = row_number  # Номер строки
        ws.cell(row=row, column=10).value = order['order_number']
        ws.cell(row=row, column=11).value = masked_phone
        ws.cell(row=row, column=12).value = masked_name
        ws.cell(row=row, column=13).value = product_name
        ws.cell(row=row, column=14).value = total_quantity
        ws.cell(row=row, column=15).value = int(order['total_amount'])
        ws.cell(row=row, column=15).number_format = '#,##0'
        
        # Применяем границы
        for col in range(9, 16):
            ws.cell(row=row, column=col).border = border
        
        row += 1
        row_number += 1
    
    # Настраиваем ширину колонок
    column_widths = {
        'A': 5,   # № (номер строки)
        'B': 12,  # Н-номер заказ
        'C': 15,  # тел-телефон
        'D': 20,  # фио
        'E': 9,   # товар (9 символов)
        'F': 12,  # я-количество
        'G': 15,  # сум-сумма заказа
        'H': 2,   # Пустая колонка между столбцами
        'I': 5,   # № (номер строки)
        'J': 12,  # Н-номер заказ
        'K': 15,  # тел-телефон
        'L': 20,  # фио
        'M': 9,   # товар (9 символов)
        'N': 12,  # я-количество
        'O': 15,  # сум-сумма заказа
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Сохраняем в байты
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    return excel_bytes


def generate_full_data_report_excel(session_id: int) -> io.BytesIO:
    """Генерирует Excel отчет с полными данными (без маскировки) в формате двух столбцов"""
    session = database.get_session(session_id)
    if not session:
        raise ValueError("Сессия не найдена")
    
    # Получаем данные
    orders = database.get_session_orders(session_id)
    
    # Создаем рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет с полными данными"
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Заголовки для первого столбца (колонки A-G)
    headers_left = ["№", "осн", "тел", "фио", "тов", "я", "сум"]
    for col, header in enumerate(headers_left, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Заголовки для второго столбца (колонки I-O)
    headers_right = ["№", "осн", "тел", "фио", "тов", "я", "сум"]
    for col, header in enumerate(headers_right, 1):
        cell = ws.cell(row=1, column=col + 8)  # +8 чтобы начать с колонки I (9-я колонка)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Разделяем заказы на две части для двух столбцов
    total_orders = len(orders)
    mid_point = (total_orders + 1) // 2  # Округляем вверх
    
    orders_left = orders[:mid_point]
    orders_right = orders[mid_point:]
    
    # Заполняем левый столбец
    row = 2
    row_number = 1
    for order in orders_left:
        # Используем полные данные без маскировки
        full_name = order['full_name'] or ""
        full_phone = order['phone_number'] or ""
        
        # Получаем товары заказа
        order_items = database.get_order_items(order['order_id'])
        if order_items:
            # Берем первый товар (или объединяем если несколько)
            product_names = [item['product_name'] for item in order_items]
            product_name = ", ".join(set(product_names))  # Убираем дубликаты
            # Обрезаем до 9 символов
            product_name = product_name[:9] if len(product_name) > 9 else product_name
            total_quantity = sum(item['quantity'] for item in order_items)
        else:
            product_name = "Нет товаров"
            total_quantity = 0
        
        # Заполняем данные
        # Используем номер по сессии, если есть, иначе порядковый номер строки
        session_num = order.get('session_order_number') or row_number
        ws.cell(row=row, column=1).value = session_num  # Номер по сессии (или порядковый номер)
        ws.cell(row=row, column=2).value = order['order_number']  # Основной номер
        ws.cell(row=row, column=3).value = full_phone
        ws.cell(row=row, column=4).value = full_name
        ws.cell(row=row, column=5).value = product_name
        ws.cell(row=row, column=6).value = total_quantity
        ws.cell(row=row, column=7).value = int(order['total_amount'])
        ws.cell(row=row, column=7).number_format = '#,##0'
        
        # Применяем границы
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        
        row += 1
        row_number += 1
    
    # Заполняем правый столбец (продолжаем нумерацию с левой таблицы)
    row = 2
    # Продолжаем нумерацию с того места, где закончилась левая таблица
    row_number = len(orders_left) + 1
    for order in orders_right:
        # Используем полные данные без маскировки
        full_name = order['full_name'] or ""
        full_phone = order['phone_number'] or ""
        
        # Получаем товары заказа
        order_items = database.get_order_items(order['order_id'])
        if order_items:
            # Берем первый товар (или объединяем если несколько)
            product_names = [item['product_name'] for item in order_items]
            product_name = ", ".join(set(product_names))  # Убираем дубликаты
            # Обрезаем до 9 символов
            product_name = product_name[:9] if len(product_name) > 9 else product_name
            total_quantity = sum(item['quantity'] for item in order_items)
        else:
            product_name = "Нет товаров"
            total_quantity = 0
        
        # Заполняем данные (начинаем с колонки I, т.е. колонка 9)
        # Используем номер по сессии, если есть, иначе порядковый номер строки
        session_num = order.get('session_order_number') or row_number
        ws.cell(row=row, column=9).value = session_num  # Номер по сессии (или порядковый номер)
        ws.cell(row=row, column=10).value = order['order_number']  # Основной номер
        ws.cell(row=row, column=11).value = full_phone
        ws.cell(row=row, column=12).value = full_name
        ws.cell(row=row, column=13).value = product_name
        ws.cell(row=row, column=14).value = total_quantity
        ws.cell(row=row, column=15).value = int(order['total_amount'])
        ws.cell(row=row, column=15).number_format = '#,##0'
        
        # Применяем границы
        for col in range(9, 16):
            ws.cell(row=row, column=col).border = border
        
        row += 1
        row_number += 1
    
    # Настраиваем ширину колонок
    column_widths = {
        'A': 5,   # № (номер строки)
        'B': 12,  # Н-номер заказ
        'C': 18,  # тел-телефон (полный номер)
        'D': 25,  # фио (полное ФИО)
        'E': 9,   # товар (9 символов)
        'F': 12,  # я-количество
        'G': 15,  # сум-сумма заказа
        'H': 2,   # Пустая колонка между столбцами
        'I': 5,   # № (номер строки)
        'J': 12,  # Н-номер заказ
        'K': 18,  # тел-телефон (полный номер)
        'L': 25,  # фио (полное ФИО)
        'M': 9,   # товар (9 символов)
        'N': 12,  # я-количество
        'O': 15,  # сум-сумма заказа
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Устанавливаем вертикальную ориентацию страницы (портретная) для формата A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9  # A4 paper size
    
    # Сохраняем в байты
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    return excel_bytes


def generate_full_data_report_html(session_id: int) -> str:
    """Генерирует HTML страницу с таблицей с полными данными (без маскировки)"""
    session = database.get_session(session_id)
    if not session:
        raise ValueError("Сессия не найдена")
    
    # Получаем данные
    orders = database.get_session_orders(session_id)
    
    # Используем полные данные без маскировки
    full_orders = []
    for order in orders:
        full_name = order['full_name'] or ""
        full_phone = order['phone_number'] or ""
        
        # Получаем товары заказа
        order_items = database.get_order_items(order['order_id'])
        if order_items:
            product_names = [item['product_name'] for item in order_items]
            product_name = ", ".join(set(product_names))
            # Обрезаем до 9 символов
            product_name = product_name[:9] if len(product_name) > 9 else product_name
            total_quantity = sum(item['quantity'] for item in order_items)
        else:
            product_name = "Нет товаров"
            total_quantity = 0
        
        full_orders.append({
            'order_number': order['order_number'],  # Основной номер
            'session_order_number': order.get('session_order_number') or '',  # Номер по сессии
            'phone': full_phone,
            'name': full_name,
            'product': product_name,
            'quantity': total_quantity,
            'amount': int(order['total_amount'])
        })
    
    # Разделяем на две части
    mid_point = (len(full_orders) + 1) // 2
    orders_left = full_orders[:mid_point]
    orders_right = full_orders[mid_point:]
    
    # Генерируем HTML
    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Полный отчет - {session['session_name']}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            display: flex;
            gap: 20px;
            justify-content: space-between;
        }}
        .table-section {{
            flex: 1;
            background-color: white;
            padding: 10px;
            border-radius: 5px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 12px;
        }}
        th {{
            background-color: #366092;
            color: white;
            padding: 8px;
            text-align: center;
            font-weight: bold;
            border: 1px solid #ddd;
        }}
        td {{
            padding: 6px;
            text-align: center;
            border: 1px solid #ddd;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        .header {{
            text-align: center;
            margin-bottom: 20px;
            font-size: 18px;
            font-weight: bold;
        }}
    </style>
</head>
<body>
    <div class="header">Полный отчет: {session['session_name']}</div>
    <div class="container">
        <div class="table-section">
            <table>
                <thead>
                    <tr>
                        <th>№</th>
                        <th>осн</th>
                        <th>тел</th>
                        <th>фио</th>
                        <th>тов</th>
                        <th>я</th>
                        <th>сум</th>
                    </tr>
                </thead>
                <tbody>
"""
    
    # Левая таблица
    row_number = 1
    for order in orders_left:
        # Используем номер по сессии, если есть, иначе порядковый номер строки
        session_num = order.get('session_order_number') or row_number
        html += f"""                    <tr>
                        <td>{session_num}</td>
                        <td>{order['order_number']}</td>
                        <td>{order['phone']}</td>
                        <td>{order['name']}</td>
                        <td>{order['product']}</td>
                        <td>{order['quantity']}</td>
                        <td>{order['amount']}</td>
                    </tr>
"""
        row_number += 1
    
    html += """                </tbody>
            </table>
        </div>
        <div class="table-section">
            <table>
                <thead>
                    <tr>
                        <th>№</th>
                        <th>осн</th>
                        <th>тел</th>
                        <th>фио</th>
                        <th>тов</th>
                        <th>я</th>
                        <th>сум</th>
                    </tr>
                </thead>
                <tbody>
"""
    
    # Правая таблица (продолжаем нумерацию с левой таблицы)
    row_number = len(orders_left) + 1
    for order in orders_right:
        # Используем номер по сессии, если есть, иначе порядковый номер строки
        session_num = order.get('session_order_number') or row_number
        html += f"""                    <tr>
                        <td>{session_num}</td>
                        <td>{order['order_number']}</td>
                        <td>{order['phone']}</td>
                        <td>{order['name']}</td>
                        <td>{order['product']}</td>
                        <td>{order['quantity']}</td>
                        <td>{order['amount']}</td>
                    </tr>
"""
        row_number += 1
    
    html += """                </tbody>
            </table>
        </div>
    </div>
</body>
</html>"""
    
    return html


def generate_channel_report_html(session_id: int) -> str:
    """Генерирует HTML страницу с таблицей для канала с маскировкой данных"""
    session = database.get_session(session_id)
    if not session:
        raise ValueError("Сессия не найдена")
    
    # Получаем данные
    orders = database.get_session_orders(session_id)
    
    # Маскируем данные
    masked_orders = []
    for order in orders:
        masked_name = qr_code.mask_name_channel(order['full_name'])
        masked_phone = qr_code.mask_phone_channel(order['phone_number'])
        
        # Получаем товары заказа
        order_items = database.get_order_items(order['order_id'])
        if order_items:
            product_names = [item['product_name'] for item in order_items]
            product_name = ", ".join(set(product_names))
            # Обрезаем до 9 символов
            product_name = product_name[:9] if len(product_name) > 9 else product_name
            total_quantity = sum(item['quantity'] for item in order_items)
        else:
            product_name = "Нет товаров"
            total_quantity = 0
        
        masked_orders.append({
            'order_number': order['order_number'],
            'phone': masked_phone,
            'name': masked_name,
            'product': product_name,
            'quantity': total_quantity,
            'amount': int(order['total_amount'])
        })
    
    # Разделяем на две части
    mid_point = (len(masked_orders) + 1) // 2
    orders_left = masked_orders[:mid_point]
    orders_right = masked_orders[mid_point:]
    
    # Генерируем HTML
    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Отчет для канала - {session['session_name']}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            display: flex;
            gap: 20px;
            justify-content: space-between;
        }}
        .table-section {{
            flex: 1;
            background-color: white;
            padding: 10px;
            border-radius: 5px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 12px;
        }}
        th {{
            background-color: #366092;
            color: white;
            padding: 8px;
            text-align: center;
            font-weight: bold;
            border: 1px solid #ddd;
        }}
        td {{
            padding: 6px;
            text-align: center;
            border: 1px solid #ddd;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        .header {{
            text-align: center;
            margin-bottom: 20px;
            font-size: 18px;
            font-weight: bold;
        }}
    </style>
</head>
<body>
    <div class="header">Отчет для канала: {session['session_name']}</div>
    <div class="container">
        <div class="table-section">
            <table>
                <thead>
                    <tr>
                        <th>№</th>
                        <th>н</th>
                        <th>тел</th>
                        <th>фио</th>
                        <th>тов</th>
                        <th>я</th>
                        <th>сум</th>
                    </tr>
                </thead>
                <tbody>
"""
    
    # Левая таблица
    row_number = 1
    for order in orders_left:
        html += f"""                    <tr>
                        <td>{row_number}</td>
                        <td>{order['order_number']}</td>
                        <td>{order['phone']}</td>
                        <td>{order['name']}</td>
                        <td>{order['product']}</td>
                        <td>{order['quantity']}</td>
                        <td>{order['amount']}</td>
                    </tr>
"""
        row_number += 1
    
    html += """                </tbody>
            </table>
        </div>
        <div class="table-section">
            <table>
                <thead>
                    <tr>
                        <th>№</th>
                        <th>н</th>
                        <th>тел</th>
                        <th>фио</th>
                        <th>тов</th>
                        <th>я</th>
                        <th>сум</th>
                    </tr>
                </thead>
                <tbody>
"""
    
    # Правая таблица (продолжаем нумерацию с левой таблицы)
    row_number = len(orders_left) + 1
    for order in orders_right:
        html += f"""                    <tr>
                        <td>{row_number}</td>
                        <td>{order['order_number']}</td>
                        <td>{order['phone']}</td>
                        <td>{order['name']}</td>
                        <td>{order['product']}</td>
                        <td>{order['quantity']}</td>
                        <td>{order['amount']}</td>
                    </tr>
"""
        row_number += 1
    
    html += """                </tbody>
            </table>
        </div>
    </div>
</body>
</html>"""
    
    return html


async def generate_channel_report_screenshot(session_id: int) -> io.BytesIO:
    """Генерирует скриншот таблицы для канала из браузера"""
    try:
        from playwright.async_api import async_playwright
        import tempfile
        import os
        
        # Генерируем HTML
        html_content = generate_channel_report_html(session_id)
        
        # Создаем временный HTML файл
        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
            f.write(html_content)
            temp_html_path = f.name
        
        try:
            # Используем Playwright для скриншота
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                page = await browser.new_page()
                
                # Загружаем HTML файл
                file_url = f"file://{temp_html_path}"
                await page.goto(file_url, wait_until='networkidle')
                
                # Делаем полный скриншот страницы
                screenshot_bytes = await page.screenshot(full_page=True, type='png')
                
                await browser.close()
            
            # Сохраняем в BytesIO
            screenshot_io = io.BytesIO(screenshot_bytes)
            screenshot_io.seek(0)
            
            return screenshot_io
        finally:
            # Удаляем временный файл
            if os.path.exists(temp_html_path):
                os.unlink(temp_html_path)
    except ImportError:
        raise ImportError("Playwright не установлен. Установите его командой: playwright install chromium")
    except Exception as e:
        raise Exception(f"Ошибка при создании скриншота: {str(e)}")


async def generate_full_data_report_screenshot(session_id: int) -> io.BytesIO:
    """Генерирует скриншот таблицы с полными данными из браузера"""
    try:
        from playwright.async_api import async_playwright
        import tempfile
        import os
        
        # Генерируем HTML
        html_content = generate_full_data_report_html(session_id)
        
        # Создаем временный HTML файл
        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
            f.write(html_content)
            temp_html_path = f.name
        
        try:
            # Используем Playwright для скриншота
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                page = await browser.new_page()
                
                # Загружаем HTML файл
                file_url = f"file://{temp_html_path}"
                await page.goto(file_url, wait_until='networkidle')
                
                # Делаем полный скриншот страницы
                screenshot_bytes = await page.screenshot(full_page=True, type='png')
                
                await browser.close()
            
            # Сохраняем в BytesIO
            screenshot_io = io.BytesIO(screenshot_bytes)
            screenshot_io.seek(0)
            
            return screenshot_io
        finally:
            # Удаляем временный файл
            if os.path.exists(temp_html_path):
                os.unlink(temp_html_path)
    except ImportError:
        raise ImportError("Playwright не установлен. Установите его командой: playwright install chromium")
    except Exception as e:
        raise Exception(f"Ошибка при создании скриншота: {str(e)}")


def generate_pending_orders_html(session_id: int) -> str:
    """Генерирует HTML страницу с таблицей не выданных заказов"""
    session = database.get_session(session_id)
    if not session:
        raise ValueError("Сессия не найдена")
    
    # Получаем все заказы сессии
    all_orders = database.get_session_orders(session_id)
    
    # Фильтруем только не выданные заказы (статус != 'completed' и != 'cancelled')
    pending_orders = [o for o in all_orders if o['status'] != 'completed' and o['status'] != 'cancelled']
    
    if not pending_orders:
        return f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Не выданные заказы - {session['session_name']}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
            text-align: center;
        }}
        .header {{
            font-size: 24px;
            font-weight: bold;
            margin: 50px 0;
        }}
    </style>
</head>
<body>
    <div class="header">Не выданные заказы: {session['session_name']}</div>
    <div style="font-size: 18px; margin-top: 50px;">Нет не выданных заказов</div>
</body>
</html>"""
    
    # Используем полные данные без маскировки
    full_orders = []
    for order in pending_orders:
        full_name = order['full_name'] or ""
        full_phone = order['phone_number'] or ""
        
        # Получаем товары заказа
        order_items = database.get_order_items(order['order_id'])
        if order_items:
            product_names = [item['product_name'] for item in order_items]
            product_name = ", ".join(set(product_names))
            # Обрезаем до 9 символов
            product_name = product_name[:9] if len(product_name) > 9 else product_name
            total_quantity = sum(item['quantity'] for item in order_items)
        else:
            product_name = "Нет товаров"
            total_quantity = 0
        
        # Используем номер по сессии, если есть, иначе порядковый номер
        session_num = order.get('session_order_number') or len(full_orders) + 1
        
        full_orders.append({
            'order_number': order['order_number'],  # Основной номер
            'session_order_number': session_num,  # Номер по сессии
            'phone': full_phone,
            'name': full_name,
            'product': product_name,
            'quantity': total_quantity,
            'amount': int(order['total_amount']),
            'status': order['status']
        })
    
    # Разделяем на две части
    mid_point = (len(full_orders) + 1) // 2
    orders_left = full_orders[:mid_point]
    orders_right = full_orders[mid_point:]
    
    # Генерируем HTML
    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Не выданные заказы - {session['session_name']}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            display: flex;
            gap: 20px;
            justify-content: space-between;
        }}
        .table-section {{
            flex: 1;
            background-color: white;
            padding: 10px;
            border-radius: 5px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 12px;
        }}
        th {{
            background-color: #366092;
            color: white;
            padding: 8px;
            text-align: center;
            font-weight: bold;
            border: 1px solid #ddd;
        }}
        td {{
            padding: 6px;
            text-align: center;
            border: 1px solid #ddd;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        .header {{
            text-align: center;
            margin-bottom: 20px;
            font-size: 18px;
            font-weight: bold;
        }}
    </style>
</head>
<body>
    <div class="header">Не выданные заказы: {session['session_name']}</div>
    <div class="container">
        <div class="table-section">
            <table>
                <thead>
                    <tr>
                        <th>№</th>
                        <th>осн</th>
                        <th>тел</th>
                        <th>фио</th>
                        <th>тов</th>
                        <th>я</th>
                        <th>сум</th>
                    </tr>
                </thead>
                <tbody>
"""
    
    # Левая таблица
    row_number = 1
    for order in orders_left:
        html += f"""                    <tr>
                        <td>{order['session_order_number']}</td>
                        <td>{order['order_number']}</td>
                        <td>{order['phone']}</td>
                        <td>{order['name']}</td>
                        <td>{order['product']}</td>
                        <td>{order['quantity']}</td>
                        <td>{order['amount']}</td>
                    </tr>
"""
        row_number += 1
    
    html += """                </tbody>
            </table>
        </div>
        <div class="table-section">
            <table>
                <thead>
                    <tr>
                        <th>№</th>
                        <th>осн</th>
                        <th>тел</th>
                        <th>фио</th>
                        <th>тов</th>
                        <th>я</th>
                        <th>сум</th>
                    </tr>
                </thead>
                <tbody>
"""
    
    # Правая таблица (продолжаем нумерацию с левой таблицы)
    row_number = len(orders_left) + 1
    for order in orders_right:
        html += f"""                    <tr>
                        <td>{order['session_order_number']}</td>
                        <td>{order['order_number']}</td>
                        <td>{order['phone']}</td>
                        <td>{order['name']}</td>
                        <td>{order['product']}</td>
                        <td>{order['quantity']}</td>
                        <td>{order['amount']}</td>
                    </tr>
"""
        row_number += 1
    
    html += """                </tbody>
            </table>
        </div>
    </div>
</body>
</html>"""
    
    return html


async def generate_pending_orders_screenshot(session_id: int) -> io.BytesIO:
    """Генерирует скриншот таблицы не выданных заказов"""
    try:
        from playwright.async_api import async_playwright
        import tempfile
        import os
        
        # Генерируем HTML
        html_content = generate_pending_orders_html(session_id)
        
        # Создаем временный HTML файл
        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
            f.write(html_content)
            temp_html_path = f.name
        
        try:
            # Используем Playwright для скриншота
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                page = await browser.new_page()
                
                # Загружаем HTML файл
                file_url = f"file://{temp_html_path}"
                await page.goto(file_url, wait_until='networkidle')
                
                # Делаем полный скриншот страницы
                screenshot_bytes = await page.screenshot(full_page=True, type='png')
                
                await browser.close()
            
            # Сохраняем в BytesIO
            screenshot_io = io.BytesIO(screenshot_bytes)
            screenshot_io.seek(0)
            
            return screenshot_io
        finally:
            # Удаляем временный файл
            if os.path.exists(temp_html_path):
                os.unlink(temp_html_path)
    except ImportError:
        raise ImportError("Playwright не установлен. Установите его командой: playwright install chromium")
    except Exception as e:
        raise Exception(f"Ошибка при создании скриншота: {str(e)}")
